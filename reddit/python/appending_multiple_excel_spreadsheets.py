import os
from openpyxl import load_workbook, Workbook


def concat_data(target_dir: str) -> list:
    headers = []
    data = []
    for file in os.listdir(target_dir):
        file_path = os.path.join(target_dir, file)
        if os.path.isdir(file_path) or os.path.splitext(file_path)[1] != ".xlsx":
            print(os.path.splitext(file_path)[1])
            continue
        wb = load_workbook(file_path)
        ws = wb.active
        if len(headers) == 0:
            headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)
    return headers, data


def write_output(target_dir: str, headers: list, data: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Concatenated Data"
    ws.append(headers)
    for row in data:
        ws.append(row)
    wb.save(os.path.join(target_dir, "process_output.xlsx"))


def main():
    while not os.path.isdir(target_dir := input("Enter directory where files are located.\n")):  # interactive prompt
        print("\nInvalid input. Please use absolute path. If using Windows, use \\\\ instead of \\.")

    headers, data = concat_data(target_dir)

    write_output(target_dir, headers, data)


if __name__ == '__main__':
    main()
