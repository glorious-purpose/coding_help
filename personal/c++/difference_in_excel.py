"""Grab changes between two versions of an excel spreadsheet file."""
from openpyxl import load_workbook, Workbook
import os
import pandas as pd

# FILE2 is the file to merge changes into.
FILE1 = "/mnt/N/Documentation/Stoughton Trailers/Stoughton Trailer LineSet Tracking.xlsx"
FILE2 = "/mnt/N/Documentation/Stoughton Trailers/Stoughton Trailer LineSet Tracking  SHAREPOINT.xlsx"
OUTPUT_FILE = "/mnt/N/Documentation/Stoughton Trailers/Stoughton Trailer LineSet Tracking-Differences.xlsx"

TABS = ["Lineset Tracking"]


def column_letter(col_num: int) -> str:
    output = ""
    col = col_num
    while col // 26 > 0:
        output += chr(ord("A") + col // 26 - 1)
        col = col % 26
    output += chr(ord("A") + col - 1)
    return output


owb = Workbook()
f1wb = load_workbook(FILE1)
f2wb = load_workbook(FILE2)
differences = {tab: [] for tab in TABS}

for tab in TABS:
    ws1 = f1wb[tab]
    ws2 = f2wb[tab]
    max_line = max(ws1.max_row, ws2.max_row)
    max_col = max(len(ws1[1]), len(ws2[1]))
    for i in range(2, max_line):
        if ws1[i] == ws2[i]:
            print(i)
            continue
        for j in range(0, max_col):
            if ws1[i][j].value != ws2[i][j].value:
                differences[tab].append((column_letter(j) + str(i), ws1[i][j].value, ws2[i][j].value))
                print([cell.value for cell in ws1[i]], [cell.value for cell in ws2[i]], i, j, sep="\n")
                input()
            if ws1[i][j].value is None and ws2[i][j].value is not None:
                ws1[i][j].value = ws2[i][j].value
            elif ws2[i][j].value is None and ws1[i][j].value is not None:
                ws2[i][j].value = ws2[i][j].value
f1wb.save(FILE1.replace(".xlsx", "-updated.xlsx"))
f2wb.save(FILE2.replace(".xlsx", "-updated.xlsx"))
default_sheet = owb.active
for tab in TABS:
    owb.create_sheet(tab)
    ows = owb[tab]
    for row in differences[tab]:
        ows.append(row)
owb.remove_sheet(default_sheet)
owb.save(OUTPUT_FILE)


# Get values from FILE1
